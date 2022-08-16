from openpyxl import load_workbook
from openpyxl import Workbook
import unicodedata
import re
from isbnlib import get_isbnlike, get_canonical_isbn, canonical, is_isbn10, is_isbn13


#This function extract URLs from a string, return a vector of URLs found
def URLsearch(stringinput):
  #regular expression
 regularex = r"(?i)\b((?:https?://|www\d{0,3}[.]|[a-z0-9.-]+[.][a-z]{2,4}/)(?:[^\s()<>]+|(([^\s()<>]+|(([^\s()<>]+)))))+(?:(([^\s()<>]+|(([^\s()<>]+))))|[^\s`!()[]{};:'\".,<>?«»“”‘’]))"
 #finding the url in passed string
 urlsrc = re.findall(regularex,stringinput)
 #return the found website url
 return [url[0] for url in urlsrc]
 

# Function: this function traverses a string looking for the occurrence of a pattern that indicates the beginning of a bibliography.
#Analyzing the pattern of bibliography records of a problem presented, it was noticed that a bibliography starts with the author's name in capital letters and a comma. The end is given by the end of the string or the occurrence of a new bibliography
#The problem is to identify and separete bibliograph from a string with bibliographs added without any care with pattern or text formating. The goal is to prepare a CSV file whit bibliograph to indexed cataloging to the subtitle.
def TextToBibliArray(biblitext):
    
    Bibliarray=[]
    a=0
    ini_temp=0
    d=[] #array to store start possition of a bibliograph
    for c in biblitext:
        a = a+1
        if unicodedata.category(c) == "Lu": #if caracter is # 'L'etter, 'u'ppercase 'Lu'
           if ini_temp==0: #if initial positial is not already stored,
               ini_temp=a #store temporarily the initial position

        if c== "," and ini_temp>0: #if an uppercase caracter was find (ini_temp>0) and the next caracter is ",", means that a start of a new bibliograph was find. Store 
            d.append(ini_temp-1) #append permanently the initital position
            ini_temp=0 #restore to find new occurrence again
        if unicodedata.category(c) != "Lu" and c!= ",": #if isn´t uppercase and isn´t ",", reset temporarily the initial position
            ini_temp=0
    d.append(len(biblitext))
    u = 1
    gojump=0
    while u <= len(d):
        if u<len(d): #previnir outofindex, já que o index é manipulado
            if d[u] - d[u-1] < 50: #distâncias de ocorrência  menor que 50 caracteres representam ocorrência de co-autoria
                gojump += 1 #se for menor que 50, salta para a próxima ocorrência
            if u+gojump <= len(d): #to avoid outofindex
                if u+gojump < len(d):
                    Bibliarray.append(biblitext[d[u-1]:d[u+gojump]])
                else: Bibliarray.append(biblitext[d[u-1]:d[u+gojump-1]])
        u += 1 + gojump #reposiciona index após salto
        gojump=0 #zera salto para próximas análises
    return Bibliarray

#String exemple
s = "ANDRADE, Maria Margarida de. Introdução à metodologia do trabalho científico: elaboração de trabalhos na graduação. 10ª. São Paulo Atlas 2012 1 recurso online ISBN 9788522478392. Disponível em https://integrada.minhabiblioteca.com.br/books/9788522478392.  CALIJURI, Maria do Carmo. Engenharia Ambiental: conceitos, tecnologias e gestão. 2. Rio de Janeiro GEN LTC 2019 1 recurso online ISBN 9788595157446.  Disponível em https://integrada.minhabiblioteca.com.br/books/9788595157446.  HALLIDAY, David. Fundamentos de física, v.1 mecânica. 10. São Paulo LTC 2016 1 recurso online ISBN 9788521632054. Disponível em https://integrada.minhabiblioteca.com.br/reader/books/9788521632054. BEER, Ferdinand. Mecânica vetorial para engenheiros: Estática, v. 1. 11. Porto Alegre AMGH 2019 1 recurso online ISBN 9788580556209.  Disponível em https://integrada.minhabiblioteca.com.br/books/9788580556209.   MIRANDA, Shirley Aparecida de. Diversidade e ações afirmativas combatendo as desigualdades sociais. São Paulo Autêntica 2010 1 recurso online ISBN 9788582178157.Disponível em https://integrada.minhabiblioteca.com.br/books/9788582178157."

#to uderstand how to work with Excel File, see: https://letscode.com.br/blog/aprenda-a-integrar-python-e-excel
caminho = 'base.xlsx'
arquivo_excel = load_workbook(caminho)
planilha1 = arquivo_excel.active
max_linha = planilha1.max_row - 3
max_coluna = planilha1.max_column
#max_linha = 16
max_coluna = 4
disccoll = 1
basicbiblicoll = 4
compbiblicoll = 5


#creating output file
output_excel_file = Workbook()
planilha_output_excel_file = output_excel_file.active
#writting header
planilha_output_excel_file.cell(row=1, column=1, value="Disciplina")
planilha_output_excel_file.cell(row=1, column=2, value="Bibliografia")
planilha_output_excel_file.cell(row=1, column=3, value="Tipo")
planilha_output_excel_file.cell(row=1, column=4, value="Url")
planilha_output_excel_file.cell(row=1, column=5, value="ISBN")

linha = 2
i=0

#for linha in range(1, max_linha): 
for linha in range(2, max_linha+1): 
    disciplina = planilha1.cell(row=linha, column=disccoll).value
    print(disciplina)
    basicbiblitxt = planilha1.cell(row=linha, column=basicbiblicoll).value
    basicbiblitxt = basicbiblitxt.replace("\n","")
    Bibliarray1 = TextToBibliArray(basicbiblitxt)
    #print(basicbiblitxt) #enable to see original text
    compbiblitxt = planilha1.cell(row=linha, column=compbiblicoll).value
    compbiblitxt = compbiblitxt.replace("\n","")
    Bibliarray2 = TextToBibliArray(compbiblitxt)
    #writting elements
    print(len(Bibliarray1))
    i=0
    while i < len(Bibliarray1):
        #print(i,"-",Bibliarray1[i])
        actualrow = planilha_output_excel_file.max_row + 1
        planilha_output_excel_file.cell(row=actualrow, column=1, value=disciplina)
        planilha_output_excel_file.cell(row=actualrow, column=2, value=Bibliarray1[i])
        planilha_output_excel_file.cell(row=actualrow, column=3, value="Básica")
        urls = URLsearch(Bibliarray1[i])
        if len(urls)>0:
            planilha_output_excel_file.cell(row=actualrow, column=4, value=urls[0])
        #to understand how to get isbn, see: https://pypi.org/project/isbnlib/
        isbn = get_canonical_isbn(Bibliarray1[i], output='bouth')
        planilha_output_excel_file.cell(row=actualrow, column=5, value=isbn)  
        planilha_output_excel_file.append
        i += 1
    j=0
    while j < len(Bibliarray2):
        actualrow = planilha_output_excel_file.max_row + 1
        planilha_output_excel_file.cell(row=actualrow, column=1, value=disciplina)
        planilha_output_excel_file.cell(row=actualrow, column=2, value=Bibliarray2[j])
        planilha_output_excel_file.cell(row=actualrow, column=3, value="Complementar")
        urls = URLsearch(Bibliarray2[j])
        if len(urls)>0:
            planilha_output_excel_file.cell(row=actualrow, column=4, value=urls[0])
        isbn = get_canonical_isbn(Bibliarray2[j], output='bouth')
        planilha_output_excel_file.cell(row=actualrow, column=5, value=isbn)  
        planilha_output_excel_file.append
        j += 1

output_excel_file.save('output.xlsx')